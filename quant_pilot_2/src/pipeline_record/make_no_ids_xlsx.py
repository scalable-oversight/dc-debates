"""Produce an anonymized copy of the cleaned XLSX with PIDs and demographics removed.

This script is preserved as a process record. It is NOT runnable from the
public distribution: the with-IDs input it reads
(`data/cleaned/quant-1-pilot-2.xlsx`) was withheld because it retains
`PROLIFIC_PID` and self-reported demographics. The de-identified output
this script produces is shipped directly at
`data/cleaned/quant-1-pilot-2-no-ids.xlsx`. See `src/pipeline_record/README.md`.

Strips PROLIFIC_PID and self-reported demographic columns so the file can be
shared without identifying participants. URL columns remain clickable.
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


DEMOGRAPHIC_COLS = {
    "PROLIFIC_PID",
    "degreeMajor", "physicsMinor", "currentMajor",
    "educationCompleted", "inSchool",
    "physics_teaching", "physics_knowledge_self",
    "computer_programming", "knows_programming", "cs_work",
    "show_physics_elaborate", "show_cs_elaborate",
    "cs_elaboration", "physics_elaboration",
    "gender", "age", "country", "englishLevel",
    "educationPursuing",
}

URL_COL_NAMES = {
    "section_1_url", "section_2_url", "section_3_url", "section_4_url",
}


def anonymize(input_xlsx: Path, output_xlsx: Path):
    if not input_xlsx.exists():
        raise FileNotFoundError(
            f"{input_xlsx} not found. The with-IDs cleaned XLSX was withheld "
            "from the public release because it retains PROLIFIC_PID and "
            "self-reported demographic columns; see src/pipeline_record/README.md. "
            "The de-identified output of this script is shipped at "
            "data/cleaned/quant-1-pilot-2-no-ids.xlsx."
        )
    df = pd.read_excel(input_xlsx)
    keep_cols = [c for c in df.columns if c not in DEMOGRAPHIC_COLS]
    df = df[keep_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    hdr_font = Font(name="Arial", bold=True, size=10)
    cell_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    wrap = Alignment(wrap_text=True, vertical="top")

    for ci, name in enumerate(keep_cols, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = hdr_font
        c.alignment = wrap

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col_name in enumerate(keep_cols, 1):
            val = row[col_name]
            if isinstance(val, float) and pd.isna(val):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = cell_font
            cell.alignment = wrap
            if (
                col_name in URL_COL_NAMES
                and isinstance(val, str)
                and val.startswith("http")
            ):
                cell.hyperlink = val
                cell.font = link_font

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    print(f"Wrote {len(df)} rows to {output_xlsx}")


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--input", type=Path, required=True,
                    help="Path to cleaned quant-1-pilot.xlsx")
    ap.add_argument("--output", type=Path, required=True,
                    help="Path to write the no-ids XLSX")
    args = ap.parse_args()
    anonymize(args.input, args.output)


if __name__ == "__main__":
    main()
