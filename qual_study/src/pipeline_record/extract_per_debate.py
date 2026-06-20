"""Per-debate extraction step (process record; not runnable in the public release).

Reads one raw GuidedTrack CSV (qual-1-debate-{n}.csv), joins it with the study
assignment table, hex-decodes the answer_* fields, derives final credence /
correctness / participant_message, computes per-section minutes, and writes a
per-debate "for grading" xlsx.

The raw CSV and the per-debate "for grading" xlsx files this script reads and
writes were withheld from the public release because they retain PROLIFIC_PID
and other identity columns. See src/pipeline_record/README.md.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_RAW_DIR = PROJECT_ROOT / "data" / "raw"
DEFAULT_LOOKUPS_DIR = PROJECT_ROOT / "data" / "lookups"
DEFAULT_OUT_DIR = PROJECT_ROOT / "data" / "cleaned" / "per_debate"

# Snapshot cutoff: drop responses finishing on/after 2026-04-08,
# when the study formally ended and analysis becan
CUTOFF_FINISH_UTC = pd.Timestamp("2026-04-08 00:00:00", tz="UTC")

DECODE_ERROR_PREFIX = (
    "[Decoding Error: non-hexadecimal number found in fromhex() arg at position 0] "
)

# Manual corrections applied to the 2026-04-08 snapshot. The 'switch' row for
# this participant has answer_4 pasted as plaintext (not hex), so the decoder
# fails and no credence is parseable. The literal PROLIFIC_PID has been replaced
# with a placeholder in the public release; the live override was applied
# upstream before the noids xlsx was written.
_MANUAL_OVERRIDES = [
    {
        "PROLIFIC_PID": "[REDACTED]",
        "debate_name": "switch",
        "final_credence_in_correct_answer": "68%",
        "participant_is_correct": "Yes",
    },
]


def _apply_manual_overrides(df: pd.DataFrame, debate_name_col: str) -> None:
    for ov in _MANUAL_OVERRIDES:
        if debate_name_col not in df.columns:
            continue
        mask = (df["PROLIFIC_PID"] == ov["PROLIFIC_PID"]) & (
            df[debate_name_col] == ov["debate_name"]
        )
        if not mask.any():
            continue
        if "answer_4" in df.columns and "other_comments" in df.columns:
            df.loc[mask, "other_comments"] = df.loc[mask, "answer_4"].map(
                lambda v: v[len(DECODE_ERROR_PREFIX):]
                if isinstance(v, str) and v.startswith(DECODE_ERROR_PREFIX)
                else v
            )
        df.loc[mask, "final_credence_in_correct_answer"] = ov[
            "final_credence_in_correct_answer"
        ]
        df.loc[mask, "participant_is_correct"] = ov["participant_is_correct"]


def decode_answer(encoded_str):
    """Decode the custom hex string back to UTF-8 text."""
    if pd.isna(encoded_str) or not isinstance(encoded_str, str):
        return encoded_str
    encoded_str = encoded_str.strip()
    if not encoded_str:
        return encoded_str
    try:
        return bytes.fromhex(encoded_str).decode("utf-8")
    except Exception as e:
        return f"[Decoding Error: {e}] {encoded_str}"


def extract_time(text):
    """Parse the 'START TIME ...' timestamp embedded in the decoded answer text."""
    if not isinstance(text, str):
        return pd.NaT
    match = re.search(r"START TIME\s+(.*)", text)
    if not match:
        return pd.NaT
    time_str = re.sub(r"\(.*?\)", "", match.group(1).strip()).strip()
    try:
        return pd.to_datetime(time_str, utc=True)
    except Exception:
        return pd.NaT


def parse_finish_time(text):
    """Parse the 'Time Finished (UTC)' column (yyyy-mm-dd HH:MM:SS)."""
    if pd.isna(text):
        return pd.NaT
    clean = str(text).replace("\xa0", " ").strip()
    try:
        return pd.to_datetime(clean, format="%Y-%m-%d %H:%M:%S", utc=True)
    except Exception:
        try:
            return pd.to_datetime(clean, utc=True)
        except Exception:
            return pd.NaT


def extract_credence(answer_text, debater_letter):
    """Pull the credence percentage assigned to a given debater from decoded answer text.

    Recognises 'Debater X: NN%', 'Debater X: >99.5%', and 'Debater X: <0.5%'.
    """
    if not isinstance(answer_text, str) or not isinstance(debater_letter, str):
        return None
    pattern = rf"Debater\s+{re.escape(debater_letter)}:\s*(<0\.5|>99\.5|\d{{1,2}})%"
    match = re.search(pattern, answer_text)
    return match.group(1) if match else None


def credence_to_numeric(credence_str):
    if credence_str is None or (isinstance(credence_str, float) and pd.isna(credence_str)):
        return None
    s = str(credence_str).strip()
    if s == ">99.5":
        return 99.75
    if s == "<0.5":
        return 0.25
    try:
        return float(s)
    except ValueError:
        return None


def classify_correctness(credence_display):
    if not isinstance(credence_display, str) or not credence_display:
        return ""
    val = credence_to_numeric(credence_display.rstrip("%"))
    if val is None:
        return ""
    if val > 50:
        return "Yes"
    if val == 50:
        return "-"
    return "No"


def build_message(row):
    credence_display = row["final_credence_in_correct_answer"]
    debater = row["correct_debater"]
    if not isinstance(credence_display, str) or not credence_display:
        return ""
    if not isinstance(debater, str) or not debater:
        return ""
    val = credence_to_numeric(credence_display.rstrip("%"))
    if val is None:
        return ""
    if val > 75:
        return f"As you concluded, Debater {debater} was correct"
    if val >= 51:
        return f"As you suspected, Debater {debater} was correct"
    if val == 50:
        return f"Debater {debater} was correct"
    return f"Debater {debater} was actually correct"


def extract_per_debate(
    debate_number: int,
    raw_dir: Path = DEFAULT_RAW_DIR,
    lookups_dir: Path = DEFAULT_LOOKUPS_DIR,
    out_dir: Path = DEFAULT_OUT_DIR,
) -> Path:
    raw_dir = Path(raw_dir)
    lookups_dir = Path(lookups_dir)
    out_dir = Path(out_dir)

    input_file = raw_dir / f"qual-1-debate-{debate_number}.csv"
    if not input_file.exists():
        raise FileNotFoundError(
            f"{input_file} not found. The raw GuidedTrack CSVs were withheld "
            "from the public release because they retain PROLIFIC_PID and "
            "other identity columns; see src/pipeline_record/README.md. The "
            "de-identified equivalent of this pipeline's output is "
            "data/cleaned/qual-1-all-debates-for-grading_noids.xlsx."
        )

    print(f"Reading: {input_file}")
    df = pd.read_csv(input_file, encoding="utf-8")

    assignment_file = lookups_dir / "study_assignment_by_participant_extended.csv"
    debate_ids_file = lookups_dir / "debate_ids.xlsx"

    debate_name_col = f"debate_{debate_number}_name"
    condition_col = f"condition_{debate_number}"
    turn4_col = f"debate_{debate_number}_turn_4"

    assignment_cols = [
        "participant_id",
        "group",
        "simplified-category",
        "category",
        debate_name_col,
        condition_col,
        "correct_debater",
        turn4_col,
    ]

    if assignment_file.exists():
        print(f"Reading: {assignment_file}")
        assign_df = pd.read_csv(assignment_file, encoding="utf-8")
        available = [c for c in assignment_cols if c in assign_df.columns]
        missing = [c for c in assignment_cols if c not in assign_df.columns and c != "correct_debater"]
        if missing:
            print(f"Warning: assignment file missing columns: {missing}")
        assign_df = assign_df[["PROLIFIC_PID", *available]].copy()
        if turn4_col in assign_df.columns:
            assign_df[turn4_col] = assign_df[turn4_col].apply(
                lambda v: v.replace(
                    "../stimuli_generation/stimuli/",
                    "https://moduloresearch.com/debates/prolific_qual_1/",
                )
                if isinstance(v, str)
                else v
            )
        df = df.merge(assign_df, on="PROLIFIC_PID", how="left")
    else:
        print(f"Warning: {assignment_file} not found; skipping assignment merge.")

    if condition_col in df.columns:
        df["correct_debater"] = df[condition_col].map(
            {"Honest first": "A", "Dishonest first": "B"}
        )
    else:
        df["correct_debater"] = ""

    requested_cols = [
        "Run",
        "Program Version",
        "guidedTrackID",
        "STUDY_ID",
        "Time Started (UTC)",
        "Time Finished (UTC)",
        "Minutes Spent",
        "Position",
        "PROLIFIC_PID",
        "pid_to_use",
        *[c for c in assignment_cols if c in df.columns],
        "device",
        "iAgreeNoLLMs",
        "consentedToAll",
        "answer_1",
        "answer_2",
        "answer_3",
        "answer_4",
        "other_comments",
    ]
    missing = [c for c in requested_cols if c not in df.columns]
    if missing:
        print(f"Warning: missing columns from raw CSV: {missing}")

    present_cols = [c for c in requested_cols if c in df.columns]
    out_df = df[present_cols].copy()

    if "Time Finished (UTC)" in out_df.columns:
        finish = out_df["Time Finished (UTC)"].apply(parse_finish_time)
        before = len(out_df)
        # Keep rows with no finish timestamp (incomplete attempts in older
        # exports) — only drop rows that demonstrably finished on/after cutoff.
        out_df = out_df[finish.isna() | (finish < CUTOFF_FINISH_UTC)].copy()
        print(
            f"Rows remaining after Time Finished < {CUTOFF_FINISH_UTC.date()} cutoff: "
            f"{len(out_df)} (dropped {before - len(out_df)})"
        )

    if "answer_4" in out_df.columns:
        out_df = out_df[
            out_df["answer_4"].apply(lambda x: isinstance(x, str) and len(x.strip()) > 15)
        ].copy()
        print(f"Rows remaining after answer_4 length filter: {len(out_df)}")

    answer_cols = ["answer_1", "answer_2", "answer_3", "answer_4"]
    for col in answer_cols:
        if col in out_df.columns:
            out_df[col] = out_df[col].apply(decode_answer)

    if "answer_4" in out_df.columns and "correct_debater" in out_df.columns:
        out_df["final_credence_in_correct_answer"] = out_df.apply(
            lambda row: extract_credence(row["answer_4"], row["correct_debater"]),
            axis=1,
        )
        out_df["final_credence_in_correct_answer"] = out_df[
            "final_credence_in_correct_answer"
        ].apply(
            lambda v: f"{v}%" if v is not None and not (isinstance(v, float) and pd.isna(v)) else ""
        )
        out_df["participant_is_correct"] = out_df["final_credence_in_correct_answer"].apply(
            classify_correctness
        )
        out_df["participant_message"] = out_df.apply(build_message, axis=1)

        if debate_ids_file.exists() and debate_name_col in out_df.columns:
            debate_ids_df = pd.read_excel(debate_ids_file)
            if {"name", "message"}.issubset(debate_ids_df.columns):
                name_to_message = dict(
                    zip(debate_ids_df["name"], debate_ids_df["message"])
                )
                out_df["participant_message"] = out_df.apply(
                    lambda row: (
                        row["participant_message"]
                        + ". "
                        + str(name_to_message[row[debate_name_col]])
                        if isinstance(row["participant_message"], str)
                        and row["participant_message"]
                        and row[debate_name_col] in name_to_message
                        and pd.notna(name_to_message[row[debate_name_col]])
                        else row["participant_message"]
                    ),
                    axis=1,
                )
            else:
                print(f"Warning: {debate_ids_file} missing 'name'/'message' columns.")
        elif not debate_ids_file.exists():
            print(f"Warning: {debate_ids_file} not found; skipping debate-specific message.")
    else:
        out_df["final_credence_in_correct_answer"] = ""
        out_df["participant_is_correct"] = ""
        out_df["participant_message"] = ""

    _apply_manual_overrides(out_df, debate_name_col)

    t1 = out_df.get("answer_1", pd.Series(dtype=object)).apply(extract_time)
    t2 = out_df.get("answer_2", pd.Series(dtype=object)).apply(extract_time)
    t3 = out_df.get("answer_3", pd.Series(dtype=object)).apply(extract_time)
    t4 = out_df.get("answer_4", pd.Series(dtype=object)).apply(extract_time)
    t_end = out_df.get("Time Finished (UTC)", pd.Series(dtype=object)).apply(parse_finish_time)

    out_df["answer_1_minutes"] = ((t2 - t1).dt.total_seconds() / 60.0).round(2)
    out_df["answer_2_minutes"] = ((t3 - t2).dt.total_seconds() / 60.0).round(2)
    out_df["answer_3_minutes"] = ((t4 - t3).dt.total_seconds() / 60.0).round(2)
    out_df["answer_4_and_comments_minutes"] = ((t_end - t4).dt.total_seconds() / 60.0).round(2)

    blank_cols = ["red_flags", "stage_1", "stage_2", "stage_3", "stage_4"]
    for col in blank_cols:
        out_df[col] = ""

    if "PROLIFIC_PID" in out_df.columns and "pid_to_use" in out_df.columns:
        mismatch = out_df["PROLIFIC_PID"].astype(str) != out_df["pid_to_use"].astype(str)
        out_df.loc[mismatch, "red_flags"] = "PROLIFIC_PID != pid_to_use"

    calc_cols = [
        "answer_1_minutes",
        "answer_2_minutes",
        "answer_3_minutes",
        "answer_4_and_comments_minutes",
    ]
    post_answer_cols = ["final_credence_in_correct_answer", "participant_is_correct"]
    extra_after_answers = ["other_comments"]
    before_answers = [
        c for c in requested_cols
        if c in out_df.columns and c not in answer_cols and c not in extra_after_answers
    ]
    final_cols = (
        before_answers
        + calc_cols
        + [c for c in answer_cols if c in out_df.columns]
        + [c for c in extra_after_answers if c in out_df.columns]
        + post_answer_cols
        + blank_cols
        + ["participant_message"]
    )
    out_df = out_df[final_cols]

    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_out = out_dir / f"qual-1-debate-{debate_number}-for-grading.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"

    header_font = Font(name="Arial", bold=True, size=10)
    cell_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(final_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = wrap

    turn4_col_idx = final_cols.index(turn4_col) + 1 if turn4_col in final_cols else None

    for row_idx, (_, row) in enumerate(out_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(final_cols, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = wrap
            if col_idx == turn4_col_idx and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = link_font

    for col_idx, col_name in enumerate(final_cols, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(ws.max_row + 1, 20)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 2, 50
        )

    wb.save(xlsx_out)
    print(f"Wrote {xlsx_out}")
    return xlsx_out


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract per-debate grading xlsx.")
    parser.add_argument("debate_number", nargs="?", default=1, type=int)
    args = parser.parse_args()
    extract_per_debate(args.debate_number)
