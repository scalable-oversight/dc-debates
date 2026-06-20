"""Split qual-1-all-debates-for-grading_noids.xlsx into one xlsx per debate_name.

Each output file contains:
  - the per-debate subset of the combined noids xlsx, with a few extra columns
    dropped (Program Version, guidedTrackID, STUDY_ID, PROLIFIC_PID,
    participant_message) on top of the columns already dropped upstream;
  - a 'Coding' sheet with columns (participant_id, answer, reasoning), where
    'answer' is 1..4 and 'reasoning' is the REASONING block extracted from
    answer_1..answer_4 respectively.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_IN_FILE = PROJECT_ROOT / "data" / "cleaned" / "qual-1-all-debates-for-grading_noids.xlsx"
DEFAULT_OUT_DIR = PROJECT_ROOT / "data" / "cleaned" / "utterance_ids"

EXTRA_DROP_COLS = (
    "Program Version",
    "guidedTrackID",
    "STUDY_ID",
    "PROLIFIC_PID",
    "participant_message",
)

DEBATE_ABBREVIATIONS = {
    "bergs": "bergs",
    "early_universe_galaxies": "eug",
    "gravitational_lens_modelling": "lensing",
    "internal_temperature_of_stars": "temp",
    "large_linear_structure": "linear",
    "little_red_dots": "lrds",
    "malt-public_306440_(rust)": "rust",
    "malt-public_323067_(db_deletion)": "dbdeletion",
    "malt-public_338159_(orm_library)": "orm",
    "malt-public_339242_(prefix-sum)": "prefixsum",
    "probability_-_expected_number_of_rolls": "prob",
    "switch": "switch",
}

REASONING_RE = re.compile(r"REASONING\s*\n(.*)", re.DOTALL)


def extract_reasoning(answer_text) -> str | None:
    if not isinstance(answer_text, str):
        return None
    match = REASONING_RE.search(answer_text)
    return match.group(1).strip() if match else None


def build_coding_df(subset: pd.DataFrame, debate_abbr: str) -> pd.DataFrame:
    rows = []
    for _, row in subset.iterrows():
        pid = row.get("participant_id")
        for n in range(1, 5):
            reasoning = extract_reasoning(row.get(f"answer_{n}"))
            if reasoning is None:
                continue
            rows.append({
                "utterance_id": f"{debate_abbr}-{pid}-{n}",
                "participant_id": pid,
                "answer": n,
                "reasoning": reasoning,
            })
    return pd.DataFrame(rows, columns=["utterance_id", "participant_id", "answer", "reasoning"])


def safe_debate_filename(debate_name: str) -> str:
    return debate_name.replace(" ", "_").replace("/", "_") + ".xlsx"


def split_by_debate(
    in_file: Path = DEFAULT_IN_FILE,
    out_dir: Path = DEFAULT_OUT_DIR,
) -> Path:
    in_file = Path(in_file)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(in_file, engine="openpyxl")
    df = df.drop(columns=[c for c in EXTRA_DROP_COLS if c in df.columns])

    debate_names = sorted(df["debate_name"].dropna().unique())
    unknown = [n for n in debate_names if n not in DEBATE_ABBREVIATIONS]
    if unknown:
        raise ValueError(f"No abbreviation for debate(s): {unknown}")

    for debate_name in debate_names:
        subset = df[df["debate_name"] == debate_name]
        if "order" in subset.columns:
            subset = subset.sort_values("order")

        coding = build_coding_df(subset, DEBATE_ABBREVIATIONS[debate_name])

        out = out_dir / safe_debate_filename(debate_name)
        subset.to_excel(out, index=False, engine="openpyxl", sheet_name="Sheet1")

        wb = load_workbook(out)
        ws = wb.create_sheet("Coding")
        for r in dataframe_to_rows(coding, index=False, header=True):
            ws.append(r)
        wb.save(out)

        print(f"  {out.name}: {len(subset)} rows, {len(coding)} coding entries")

    print(f"\nWrote {len(debate_names)} files to {out_dir}")
    return out_dir


if __name__ == "__main__":
    split_by_debate()
