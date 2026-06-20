"""Process the raw GuidedTrack export into the cleaned XLSX + warnings file.

This script is preserved as a process record. It is NOT runnable from the
public distribution: the raw inputs were withheld because they retain
`PROLIFIC_PID` and self-reported demographics, and the with-IDs output and
warnings file it would write retain the same identifiers. See
`src/pipeline_record/README.md`.

Withheld inputs:
  - `data/raw/quant-1-official-study.csv` — raw GuidedTrack export.
  - `data/raw/ai_tasker_demographics/*.csv` — Prolific demographic exports
    whose union of `Participant id` columns defines the `ai_tasker` flag.

Withheld outputs:
  - `data/cleaned/quant-1-official-study.xlsx` — with-IDs cleaned dataset.
  - `data/cleaned/warnings.txt` — warning lines quote PROLIFIC_PID values.

The de-identified equivalent of this pipeline's output is shipped at
`data/cleaned/quant-1-official-study-noids.xlsx`.

Drops duplicate question-text columns, decodes hex-encoded answers (with
lenient handling for whitespace-corrupted and partially-marked payloads),
extracts timing and credence data, merges assignment/condition metadata,
derives a per-row `sourcetype` from the raw `assignments_to_sourcetypes`
JSON column, flags AI-tasker participants from a directory of Prolific
demographic exports, flags master's/doctorate participants from the
`educationCompleted` column, splits rows into accepted/rejected based on
per-row rejection reasons (the rejected frame is not written to disk but
is used internally to suppress duplicate-PID warnings for participants
with at least one accepted submission), and writes a formatted XLSX with
clickable URL columns.
"""

import argparse
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


MASTERS_EDUCATION_LEVELS = {
    "Graduate degree (MA, MSc, MPhil, or equivalent)",
    "Doctorate degree (PhD, DPhil, or equivalent)",
}


# ── Decoding ────────────────────────────────────────────────────────────────

ENCODING_MAP = list("0123456789aBcDeF")
# Accept both cases of every hex digit so payloads whose case got mangled in
# transit (e.g. uppercase 'E' instead of the canonical lowercase 'e') still
# decode. The encoded alphabet uses fixed alternating case for obfuscation,
# but case carries no information — both cases are valid hex digits.
REVERSE_MAP = {}
for _i, _ch in enumerate(ENCODING_MAP):
    REVERSE_MAP[_ch.lower()] = _i
    REVERSE_MAP[_ch.upper()] = _i


def decode_payload(encoded):
    raw = bytearray()
    for i in range(0, len(encoded) - 1, 2):
        hi, lo = REVERSE_MAP.get(encoded[i]), REVERSE_MAP.get(encoded[i + 1])
        if hi is None or lo is None:
            return f"[DECODE ERROR at pos {i}: {encoded[i:i+2]!r}]"
        raw.append((hi << 4) | lo)
    return raw.decode("utf-8", errors="replace")


def decode_hex(encoded):
    try:
        return bytes.fromhex(encoded).decode("utf-8")
    except Exception:
        return None


def decode_answer(cell, embedded_log, run=None, col=None):
    if not isinstance(cell, str) or not cell.strip():
        return ""
    cell = cell.strip()
    m = re.match(r"^\[ROUND\d+\](.*)\[ROUND\d+\]$", cell, re.DOTALL)
    if m:
        inner = m.group(1)
    else:
        # Lenient fallback: some submissions arrive with the opening [ROUNDN]
        # marker present but the closing marker missing (probably truncated on
        # resubmission). Accept the payload as long as it starts with a marker.
        m_open = re.match(r"^\[ROUND\d+\](.*)$", cell, re.DOTALL)
        if not m_open:
            return ""
        inner = m_open.group(1)
    # Resubmissions can produce adjacent [ROUNDN][ROUNDN] marker pairs inside
    # the outer wrapper. Split on those markers and decode each hex chunk
    # separately, joining with newlines so START/END TIME lines stay parseable.
    chunks = [c for c in re.split(r"\[ROUND\d+\]", inner) if c]
    decoded = "\n".join(decode_payload(c) for c in chunks)
    m2 = re.match(r"^\[ROUND\d+\](.*)\[ROUND\d+\]$", decoded.strip(), re.DOTALL)
    if m2:
        decoded = decode_payload(m2.group(1))
    # A few answers had an entire encoded block pasted into a free-text field;
    # if we find one and it decodes to something that looks like a real answer,
    # use the inner block instead.
    m3 = re.search(r"\[ROUND\d+\]([\da-fA-F]+)\[ROUND\d+\]", decoded)
    if m3:
        inner_block = decode_hex(m3.group(1))
        if inner_block and "START TIME" in inner_block:
            embedded_log.append((run, col))
            decoded = inner_block
    return decoded


# ── Helpers ──────────────────────────────────────────────────────────────────


def is_valid_pid(pid):
    return bool(isinstance(pid, str) and re.fullmatch(r"[0-9a-f]{24}", pid))


def extract_times(text):
    def _parse(label):
        if not isinstance(text, str):
            return None
        match = re.search(rf"{label}\n(.*)", text)
        if not match:
            return None
        ts = re.sub(r"\(.*?\)", "", match.group(1)).strip()
        try:
            return pd.to_datetime(ts, utc=True)
        except Exception:
            return None
    return _parse("START TIME"), _parse("END TIME")


def extract_credence(answer_text, debater_letter):
    if not isinstance(answer_text, str) or not isinstance(debater_letter, str):
        return np.nan
    m = re.search(
        rf"Debater\s+{re.escape(debater_letter)}:\s*(\d{{1,2}})%",
        answer_text,
    )
    if not m:
        return np.nan
    return int(m.group(1)) / 100.0


def to_logit(p):
    if pd.isna(p):
        return np.nan
    p_clamped = max(0.01, min(0.99, p))
    return np.log(p_clamped / (1.0 - p_clamped))


def load_ai_tasker_pids(directory: Path):
    pids = set()
    for path in sorted(directory.glob("*.csv")):
        demo = pd.read_csv(path, encoding="utf-8", dtype=str)
        pids.update(demo["Participant id"].dropna().tolist())
    return pids


# ── Columns to keep (drop duplicate question-text columns) ──────────────────

KEEP_COLS = [
    "Run", "Program Version", "User",
    "Time Started (UTC)", "Time Finished (UTC)", "Minutes Spent",
    "Position", "Points",
    "PROLIFIC_PID", "assignment",
    "comp1", "comp2", "comp3", "comp4",
    "watched_demo_video",
    "turn1_urls", "turn2_urls", "turn3_urls", "turn4_urls",
    "guidedTrackID",
    "answer_1", "answer_2", "answer_3", "answer_4",
    "iAgreeNoLLMs", "device",
    "consentInfoSheet", "consentVoluntary", "consentDataStorage",
    "consentAge", "consentDataProcessing", "consentToParticipate",
    "consentedToAll", "confirmNoConsent",
    "degreeMajor", "physicsMinor", "currentMajor",
    "educationCompleted", "inSchool",
    "physics_teaching", "physics_knowledge_self",
    "computer_programming", "knows_programming", "cs_work",
    "show_physics_elaborate", "show_cs_elaborate",
    "cs_elaboration", "physics_elaboration",
    "gender", "age", "country", "englishLevel",
    "educationPursuing",
    "answer1check", "answer2check", "answer3check", "answer4check",
    "other_comments",
]


def process(
    input_csv: Path,
    lookups_dir: Path,
    ai_tasker_dir: Path,
    output_xlsx: Path,
    warnings_path: Path,
):
    if not input_csv.exists():
        raise FileNotFoundError(
            f"{input_csv} not found. The raw GuidedTrack CSV was withheld from "
            "the public release because it retains PROLIFIC_PID and other "
            "identity columns; see src/pipeline_record/README.md. The "
            "de-identified equivalent of this pipeline's output is "
            "data/cleaned/quant-1-official-study-noids.xlsx."
        )
    if not ai_tasker_dir.exists():
        raise FileNotFoundError(
            f"{ai_tasker_dir} not found. The Prolific demographic exports used "
            "to derive the `ai_tasker` flag were withheld from the public "
            "release because they map PROLIFIC_PID to demographic categories; "
            "see src/pipeline_record/README.md. The de-identified equivalent "
            "of this pipeline's output (including the precomputed `ai_tasker` "
            "column) is data/cleaned/quant-1-official-study-noids.xlsx."
        )
    df_full = pd.read_csv(input_csv, encoding="utf-8", dtype=str)
    df = df_full[[c for c in KEEP_COLS if c in df_full.columns]].copy()

    # Rows where iAgreeNoLLMs == "No" are not dropped here; they get a
    # rejection reason later and are excluded from `out` before writing.

    if "assignments_to_sourcetypes" in df_full.columns:
        df["_assignments_to_sourcetypes"] = df_full["assignments_to_sourcetypes"]
    else:
        df["_assignments_to_sourcetypes"] = ""

    # Strip spaces and line breaks from the raw answer cells before any further
    # processing — incoming submissions occasionally have whitespace injected into
    # the hex payload, which would corrupt decoding and the ROUND-marker checks.
    for col in ("answer_1", "answer_2", "answer_3", "answer_4"):
        df[col] = df[col].apply(
            lambda v: re.sub(r"[ \r\n]", "", v) if isinstance(v, str) else v
        )

    for col in ("answer_1", "answer_2", "answer_3", "answer_4"):
        df[f"_raw_{col}"] = df[col]

    embedded_log = []
    for col in ("answer_1", "answer_2", "answer_3", "answer_4"):
        df[col] = df.apply(
            lambda row, c=col: decode_answer(row[c], embedded_log, run=row["Run"], col=c),
            axis=1,
        )

    if embedded_log:
        print(f"Replaced {len(embedded_log)} embedded encoded block(s):")
        for run, col in embedded_log:
            print(f"  Run {run}, {col}")
    else:
        print("No embedded encoded blocks found.")

    # Timing
    for section in range(1, 5):
        parsed = df[f"answer_{section}"].apply(extract_times)
        starts = parsed.apply(lambda t: t[0])
        ends = parsed.apply(lambda t: t[1])
        df[f"answer_{section}_minutes"] = (
            (ends - starts).dt.total_seconds() / 60.0
        ).round(2)

    df["total_minutes_evaluating_debates"] = sum(
        df[f"answer_{s}_minutes"].fillna(0) for s in range(1, 5)
    ).round(2)

    # Lookups
    mapping = pd.read_csv(
        lookups_dir / "assignment-condition-turn-mapping.csv",
        dtype=str, encoding="utf-8",
    )
    hash_lookup = {r["Hash"]: r for _, r in mapping.iterrows()}

    name_map = pd.read_csv(
        lookups_dir / "debate-to-name-and-sourcetype-mapping.csv", encoding="utf-8",
    )
    id_to_name = {
        str(int(r["index"])): re.sub(r"\.json$", "", str(r["json"]))
        for _, r in name_map.iterrows()
    }

    # Warnings (computed only for rows that pass iAgreeNoLLMs).
    df_w = (
        df.loc[df["iAgreeNoLLMs"] != "No"] if "iAgreeNoLLMs" in df.columns else df
    )

    warnings_list = []

    for _, row in df_w.iterrows():
        pid = str(row["PROLIFIC_PID"]) if pd.notna(row["PROLIFIC_PID"]) else ""
        if len(pid) >= 16 and not is_valid_pid(pid):
            warnings_list.append(
                f"Run {row['Run']}: PROLIFIC_PID '{pid}' is 16+ characters "
                f"but is not a valid 24-character lowercase-hex PID"
            )

    # Duplicate PIDs — deferred until after the accept/reject split.

    flagged_round1_4 = set()
    for _, row in df_w.iterrows():
        pid = str(row["PROLIFIC_PID"]) if pd.notna(row["PROLIFIC_PID"]) else ""
        if not is_valid_pid(pid):
            continue
        raw1 = str(row["_raw_answer_1"])
        raw4 = str(row["_raw_answer_4"])
        missing = []
        if "ROUND1" not in raw1:
            missing.append("answer_1 does not contain 'ROUND1'")
        if "ROUND4" not in raw4:
            missing.append("answer_4 does not contain 'ROUND4'")
        if missing:
            warnings_list.append(
                f"Run {row['Run']}: Valid PID but {'; '.join(missing)}"
            )
            flagged_round1_4.add(row["Run"])

    for _, row in df_w.iterrows():
        if row["Run"] in flagged_round1_4:
            continue
        pid = str(row["PROLIFIC_PID"]) if pd.notna(row["PROLIFIC_PID"]) else ""
        if not is_valid_pid(pid):
            continue
        raw2 = str(row["_raw_answer_2"])
        raw3 = str(row["_raw_answer_3"])
        missing = []
        if "ROUND2" not in raw2:
            missing.append("answer_2 does not contain 'ROUND2'")
        if "ROUND3" not in raw3:
            missing.append("answer_3 does not contain 'ROUND3'")
        if missing:
            warnings_list.append(
                f"Run {row['Run']}: {'; '.join(missing)}"
            )

    for _, row in df_w.iterrows():
        if row["Run"] in flagged_round1_4:
            continue
        pid = str(row["PROLIFIC_PID"]) if pd.notna(row["PROLIFIC_PID"]) else ""
        if not is_valid_pid(pid):
            continue
        total = row["total_minutes_evaluating_debates"]
        if pd.notna(total) and total <= 9:
            warnings_list.append(
                f"Run {row['Run']}: total_minutes_evaluating_debates "
                f"({total:.2f}) is 9 minutes or less"
            )

    for _, row in df_w.iterrows():
        assignment = str(row["assignment"]) if pd.notna(row["assignment"]) else ""
        if assignment and assignment not in hash_lookup:
            warnings_list.append(
                f"Run {row['Run']}: assignment '{assignment}' does not match "
                f"any Hash in assignment-condition-turn-mapping.csv"
            )

    # Valid PID filtered out of the xlsx: list PROLIFIC_PID and every reason
    for _, row in df_w.iterrows():
        pid = str(row["PROLIFIC_PID"]) if pd.notna(row["PROLIFIC_PID"]) else ""
        if not is_valid_pid(pid):
            continue
        reasons = []
        if "ROUND1" not in str(row["_raw_answer_1"]):
            reasons.append("answer_1 does not contain 'ROUND1'")
        if "ROUND4" not in str(row["_raw_answer_4"]):
            reasons.append("answer_4 does not contain 'ROUND4'")
        total = row["total_minutes_evaluating_debates"]
        if pd.notna(total) and total <= 9:
            reasons.append(
                f"total_minutes_evaluating_debates ({total:.2f}) "
                f"is 9 minutes or less"
            )
        if reasons:
            warnings_list.append(
                f"Run {row['Run']}: PROLIFIC_PID '{pid}' was filtered out — "
                f"{'; '.join(reasons)}"
            )

    # Enrich every row with mapping data (rejected rows too — needed for the
    # deferred duplicate-PID warning, which references the rejected frame).
    def lookup(assignment, field):
        r = hash_lookup.get(str(assignment))
        return r[field] if r is not None else ""

    df["condition"] = df["assignment"].apply(lambda a: lookup(a, "Condition"))
    df["section_1_url"] = df["assignment"].apply(lambda a: lookup(a, "turn1"))
    df["section_2_url"] = df["assignment"].apply(lambda a: lookup(a, "turn2"))
    df["section_3_url"] = df["assignment"].apply(lambda a: lookup(a, "turn3"))
    df["section_4_url"] = df["assignment"].apply(lambda a: lookup(a, "turn4"))

    def _lookup_sourcetype(row):
        raw = row.get("_assignments_to_sourcetypes", "")
        assignment = row.get("assignment", "")
        if not isinstance(raw, str) or not raw or not assignment:
            return ""
        try:
            return json.loads(raw).get(str(assignment), "")
        except (ValueError, TypeError):
            return ""

    df["sourcetype"] = df.apply(_lookup_sourcetype, axis=1)
    df = df.drop(columns=["_assignments_to_sourcetypes"])

    # Derived columns
    df["debate_id"] = df["condition"].apply(
        lambda c: c.split("-")[0] if "-" in c else ""
    )
    df["correct_debater"] = df["condition"].apply(
        lambda c: c.split("-")[1] if "-" in c else ""
    )
    df["debate_name"] = df["debate_id"].apply(
        lambda d: id_to_name.get(str(int(d)), "") if d else ""
    )

    for s in range(1, 5):
        df[f"section_{s}_credence_in_correct_answer"] = df.apply(
            lambda r, s=s: extract_credence(r[f"answer_{s}"], r["correct_debater"]),
            axis=1,
        )

    def _leans_toward(row):
        c = row["section_4_credence_in_correct_answer"]
        correct = row["correct_debater"]
        if pd.isna(c):
            return ""
        if c > 0.5:
            return correct
        if c < 0.5:
            return "A" if correct == "B" else "B"
        return "-"

    df["participant_leans_toward"] = df.apply(_leans_toward, axis=1)

    df["participant_is_correct"] = df[
        "section_4_credence_in_correct_answer"
    ].apply(
        lambda c: ("Yes" if c > 0.5 else ("No" if c < 0.5 else "-"))
        if pd.notna(c)
        else ""
    )

    for s in range(1, 5):
        col = f"section_{s}_credence_in_correct_answer"
        df[f"logit_{col}"] = df[col].apply(to_logit)

    df["credence_change"] = (
        df["section_4_credence_in_correct_answer"]
        - df["section_1_credence_in_correct_answer"]
    )
    df["credence_logits_change"] = (
        df["logit_section_4_credence_in_correct_answer"]
        - df["logit_section_1_credence_in_correct_answer"]
    )

    # Per-row rejection reasons
    def _rejection_reasons(row):
        reasons = []
        iAgree = row.get("iAgreeNoLLMs", "")
        if isinstance(iAgree, str) and iAgree == "No":
            reasons.append("iAgreeNoLLMs is 'No'")
        pid = str(row["PROLIFIC_PID"]) if pd.notna(row["PROLIFIC_PID"]) else ""
        if not is_valid_pid(pid):
            reasons.append(
                "PROLIFIC_PID is not a valid 24-character lowercase-hex PID"
            )
        if "ROUND1" not in str(row["_raw_answer_1"]):
            reasons.append("answer_1 does not contain 'ROUND1'")
        if "ROUND4" not in str(row["_raw_answer_4"]):
            reasons.append("answer_4 does not contain 'ROUND4'")
        total = row["total_minutes_evaluating_debates"]
        if pd.isna(total):
            reasons.append("total_minutes_evaluating_debates is missing")
        elif total <= 9:
            reasons.append(
                f"total_minutes_evaluating_debates ({total:.2f}) "
                f"is 9 minutes or less"
            )
        if pd.isna(row["section_1_credence_in_correct_answer"]):
            reasons.append("section_1_credence_in_correct_answer is missing")
        if pd.isna(row["section_4_credence_in_correct_answer"]):
            reasons.append("section_4_credence_in_correct_answer is missing")
        return reasons

    df["_rejection_reasons"] = df.apply(_rejection_reasons, axis=1)

    # Warn about rows that would have passed every earlier check but fail due
    # to a missing credence. Preserves the prior warning text.
    keep_mask = (
        df["PROLIFIC_PID"].apply(
            lambda p: is_valid_pid(str(p)) if pd.notna(p) else False
        )
        & df["_raw_answer_1"].apply(lambda v: "ROUND1" in str(v))
        & df["_raw_answer_4"].apply(lambda v: "ROUND4" in str(v))
        & (df["total_minutes_evaluating_debates"] > 9)
        & (df.get("iAgreeNoLLMs", pd.Series("", index=df.index)) != "No")
    )
    missing_credence_mask = (
        df["section_1_credence_in_correct_answer"].isna()
        | df["section_4_credence_in_correct_answer"].isna()
    )
    for _, row in df.loc[keep_mask & missing_credence_mask].iterrows():
        reasons = []
        if pd.isna(row["section_1_credence_in_correct_answer"]):
            reasons.append("section_1_credence_in_correct_answer is missing")
        if pd.isna(row["section_4_credence_in_correct_answer"]):
            reasons.append("section_4_credence_in_correct_answer is missing")
        warnings_list.append(
            f"Run {row['Run']}: PROLIFIC_PID '{row['PROLIFIC_PID']}' was filtered "
            f"out — {'; '.join(reasons)}"
        )

    df = df.drop(columns=[c for c in df.columns if c.startswith("_raw_")])

    # Split into accepted (out) and rejected (kept in memory only)
    is_rejected = df["_rejection_reasons"].apply(bool)
    out = df.loc[~is_rejected].drop(columns=["_rejection_reasons"]).copy()
    rejects = df.loc[is_rejected].copy()
    rejects = rejects.drop(columns=["_rejection_reasons"])

    # If a participant has at least one accepted submission, drop their other
    # (failing) submissions from the rejects pool so the dedup warning below
    # doesn't fire for them.
    successful_pids = set(out["PROLIFIC_PID"].dropna().astype(str))
    rejects = rejects.loc[
        ~rejects["PROLIFIC_PID"].astype(str).isin(successful_pids)
    ].copy()

    # Deferred duplicate-PID warning: fires only when a participant has
    # multiple valid-PID submissions and none of them were accepted.
    _dup_pool = pd.concat([
        out[["Run", "PROLIFIC_PID"]],
        rejects[["Run", "PROLIFIC_PID"]],
    ], ignore_index=True)
    _dup_pool = _dup_pool[
        _dup_pool["PROLIFIC_PID"].apply(
            lambda p: is_valid_pid(str(p)) if pd.notna(p) else False
        )
    ]
    _pid_counts = _dup_pool["PROLIFIC_PID"].value_counts()
    for pid, count in _pid_counts.items():
        if count > 1 and pid not in successful_pids:
            for run_val in _dup_pool.loc[
                _dup_pool["PROLIFIC_PID"] == pid, "Run"
            ].tolist():
                warnings_list.append(
                    f"Run {run_val}: Duplicate PROLIFIC_PID '{pid}'"
                )

    # AI tasker and masters/doctorate flags
    ai_pids = load_ai_tasker_pids(ai_tasker_dir)
    out["ai_tasker"] = out["PROLIFIC_PID"].isin(ai_pids)
    out["masters_or_doctorate"] = out["educationCompleted"].isin(
        MASTERS_EDUCATION_LEVELS
    )

    # Reorder: answer_1..4 after section_4_url; sourcetype after assignment.
    answer_cols = ["answer_1", "answer_2", "answer_3", "answer_4"]
    cols_ordered = [c for c in out.columns if c not in answer_cols]
    insert_at = cols_ordered.index("section_4_url") + 1
    for i, ac in enumerate(answer_cols):
        cols_ordered.insert(insert_at + i, ac)
    if "sourcetype" in cols_ordered and "assignment" in cols_ordered:
        cols_ordered.remove("sourcetype")
        cols_ordered.insert(cols_ordered.index("assignment") + 1, "sourcetype")
    out = out[cols_ordered]

    # Write warnings
    warnings_path.parent.mkdir(parents=True, exist_ok=True)
    with open(warnings_path, "w", encoding="utf-8") as f:
        for w in warnings_list:
            f.write(w + "\n")
    print(f"Wrote {len(warnings_list)} warning(s) to {warnings_path}")

    # Write XLSX
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(out, output_xlsx)
    print(f"Wrote {len(out)} rows to {output_xlsx}")


# openpyxl rejects ASCII control characters other than \t \n \r in cell
# values. Corrupted submissions can produce such bytes after decoding, so we
# strip them before writing.
_XLSX_ILLEGAL_RE = re.compile(r"[\000-\010\013-\014\016-\037]")


def _xlsx_safe(val):
    if isinstance(val, str):
        return _XLSX_ILLEGAL_RE.sub("", val)
    return val


def write_xlsx(out: pd.DataFrame, path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    cols = list(out.columns)
    url_col_names = {
        "section_1_url", "section_2_url", "section_3_url", "section_4_url",
    }

    hdr_font = Font(name="Arial", bold=True, size=10)
    cell_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    wrap = Alignment(wrap_text=True, vertical="top")

    for ci, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = hdr_font
        c.alignment = wrap

    for ri, (_, row) in enumerate(out.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = row[col_name]
            if isinstance(val, float) and pd.isna(val):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=_xlsx_safe(val))
            cell.font = cell_font
            cell.alignment = wrap
            if (
                col_name in url_col_names
                and isinstance(val, str)
                and val.startswith("http")
            ):
                cell.hyperlink = val
                cell.font = link_font

    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--input", type=Path, required=True,
                    help="Path to raw quant-1-official-study.csv")
    ap.add_argument("--lookups", type=Path, required=True,
                    help="Directory containing the assignment/debate mapping CSVs")
    ap.add_argument("--ai-tasker-demographics-dir", type=Path, required=True,
                    help="Directory of Prolific demographic CSV exports whose "
                         "Participant id columns are unioned to flag ai_tasker=True")
    ap.add_argument("--output", type=Path, required=True,
                    help="Path to write the cleaned XLSX")
    ap.add_argument("--warnings", type=Path, required=True,
                    help="Path to write warnings.txt")
    args = ap.parse_args()
    process(
        args.input,
        args.lookups,
        args.ai_tasker_demographics_dir,
        args.output,
        args.warnings,
    )


if __name__ == "__main__":
    main()
